"""Build SST Team Sprint Planning Tracker Excel workbook."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

OUT = Path(__file__).resolve().parents[1] / "docs" / "12-planning" / "SST_Sprint_Planning_Tracker.xlsx"

# Theme
NAVY = "1B3A4B"
TEAL = "2A6F7A"
SLATE = "4A5568"
LIGHT = "F4F7F8"
WHITE = "FFFFFF"
SOFT = "E8EEF0"
ACCENT = "C45C26"
GREEN = "1F7A4C"
AMBER = "B7791F"
RED = "B33A3A"
ROW_ALT = "F8FBFC"

thin = Border(
    left=Side(style="thin", color="D0D7DE"),
    right=Side(style="thin", color="D0D7DE"),
    top=Side(style="thin", color="D0D7DE"),
    bottom=Side(style="thin", color="D0D7DE"),
)
thick_outer = Border(
    left=Side(style="medium", color=NAVY),
    right=Side(style="medium", color=NAVY),
    top=Side(style="medium", color=NAVY),
    bottom=Side(style="medium", color=NAVY),
)


def fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def font(bold=False, size=11, color=SLATE, name="Calibri"):
    return Font(name=name, bold=bold, size=size, color=color)


def set_widths(ws, widths: dict[str, float]):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def style_header_row(ws, row: int, start_col: int, end_col: int):
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill(NAVY)
        cell.font = font(bold=True, size=11, color=WHITE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin


def style_body(ws, start_row: int, end_row: int, start_col: int, end_col: int):
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = thin
            cell.font = font(size=10)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if r % 2 == 0:
                if not cell.fill or cell.fill.fgColor is None or cell.fill.fgColor.rgb in ("00000000", None):
                    cell.fill = fill(ROW_ALT)


# ---------------------------------------------------------------------------
# Data
# ---------------------------------------------------------------------------

OWNERSHIP = [
    ["Sales", "Ananya", "E3 Requirements", "Create req → assign TA → status / SLA"],
    ["Talent Acquisition (TA)", "Mohit & Rohit", "E4 Candidates (+ offer handoff)", "Add candidate → stages → select → offer"],
    ["HR", "Shalu & Harsh", "E6 Onboarding (+ offer accept)", "Accept offer → docs / BGV → Joined"],
    ["Admin", "Harsh", "E1 / E2 / E8 Auth, Masters, Audit, Import", "Users, dropdowns, audit, CSV import"],
    ["Dashboard", "Saleena & Akash", "E7 Dashboard", "KPI cards, filters, RAG / stage views"],
    ["Review", "Akash & Gaurav", "Cross-cutting QA", "Demo checklist, bugs, acceptance"],
]

MILESTONES = [
    ["M1 Foundations", "S1–S2", "Login, masters, requirements", "Login as roles; create a requirement"],
    ["M2 Pipeline", "S3–S4", "Candidates + selection", "Full TA pipeline on one req"],
    ["M3 Close the loop", "S5–S6", "Offers → onboarding → join", "Candidate selected → joined"],
    ["M4 Insights & harden", "S7–S8", "Dashboard, import, polish", "Live KPIs + Excel import"],
]

# Sprint | Step | Owner | Module | Track | Deliverable | Demo line | Epic/FR | Status | % Done | Notes | PR Link
STEPS = [
    ["S1", "A-1", "Harsh", "Admin", "Admin", "Login + JWT session + role-based route guards", "Secure access", "E1 / FR-AUTH", "Not Started", 0, "", ""],
    ["S1", "A-2", "Harsh", "Admin", "Admin", "Seed + serve Setup Lists (stages, statuses, priorities…)", "Dropdowns match Excel", "E2 / FR-MD", "Not Started", 0, "", ""],
    ["S1", "S-1", "Ananya", "Sales", "Requirements", "Requirements list page (table, search, basic filters)", "Here's all open demand", "E3-S1 / FR-REQ", "Not Started", 0, "", ""],
    ["S1", "S-2", "Ananya", "Sales", "Requirements", "Create requirement form (client, role, positions, sales owner, priority)", "New req in 1 minute", "E3-S1 / FR-REQ-01", "Not Started", 0, "", ""],
    ["S1", "D-S1", "Saleena", "Dashboard", "Dashboard", "Dashboard shell + loading skeletons", "Home page ready", "E7", "Not Started", 0, "", ""],
    ["S1", "R-G1", "Gaurav", "Review", "Review", "Test accounts per role (Admin/Sales/TA/HR/Lead) + login matrix", "RBAC demo prep", "QA", "Not Started", 0, "", ""],
    ["S1", "R-A1", "Akash", "Review", "Review", "Sprint demo script + acceptance checklist (ongoing)", "Manager walkthrough ready", "QA", "Not Started", 0, "", ""],
    ["S2", "A-3", "Harsh", "Admin", "Admin", "Admin Users CRUD (create/disable, assign roles)", "Provision Sales/TA/HR", "E1-S3 / FR-AUTH-04", "Not Started", 0, "", ""],
    ["S2", "A-4", "Harsh", "Admin", "Admin", "Clients + Job Families CRUD", "Normalized masters", "E2-S2 / FR-MD-08/09", "Not Started", 0, "", ""],
    ["S2", "S-3", "Ananya", "Sales", "Requirements", "Requirement detail page + edit", "Edit without breaking Excel rules", "E3 / FR-REQ", "Not Started", 0, "", ""],
    ["S2", "S-4", "Ananya", "Sales", "Requirements", "Assign TA + TA handoff date", "Sales → TA handoff", "E3-S2 / FR-REQ-02", "Not Started", 0, "", ""],
    ["S2", "T-M1", "Mohit", "TA", "Candidates", "Candidates list + filter by requirement", "Pipeline inventory", "E4 / FR-CAN-10", "Not Started", 0, "", ""],
    ["S2", "D-S1b", "Saleena", "Dashboard", "Dashboard", "KPI card layout shell (zeros OK)", "Demand cards placeholder", "E7-S1", "Not Started", 0, "", ""],
    ["S2", "R-G2", "Gaurav", "Review", "Review", "Requirements acceptance checklist vs FR-REQ", "Sales module sign-off", "QA", "Not Started", 0, "", ""],
    ["S3", "S-5", "Ananya", "Sales", "Requirements", "Status transitions: Active / On Hold / Cancelled / Closed", "Lifecycle controls", "E3-S3 / FR-REQ-08", "Not Started", 0, "", ""],
    ["S3", "S-6", "Ananya", "Sales", "Requirements", "Derived fields: age, open/closed positions, TA handoff RAG", "SLA visible without Excel formulas", "E3-S4 / FR-REQ-05/06/07", "Not Started", 0, "", ""],
    ["S3", "T-M2", "Mohit", "TA", "Candidates", "Add candidate from requirement detail (prefill client/role/owners)", "Add profile in context", "E4-S1 / FR-CAN-01/02/09", "Not Started", 0, "", ""],
    ["S3", "T-M3", "Mohit", "TA", "Candidates", "Stage + feedback updates (dropdowns from masters)", "Stage moves like Excel, but audited", "E4-S2 / FR-CAN-03", "Not Started", 0, "", ""],
    ["S3", "T-R1", "Rohit", "TA", "Candidates", "Candidate form validation + link integrity to Req ID", "Data quality at entry", "E4 / FR-CAN", "Not Started", 0, "", ""],
    ["S3", "A-5", "Harsh", "Admin", "Admin", "Wire all forms to master-data APIs", "No hard-coded lists", "E2-S3", "Not Started", 0, "", ""],
    ["S3", "D-S2", "Saleena", "Dashboard", "Dashboard", "KPI cards: total reqs, positions, open/closed", "Demand snapshot", "E7-S1 / FR-DASH-01", "Not Started", 0, "", ""],
    ["S3", "D-A1", "Akash", "Dashboard", "Dashboard", "Filter bar: TA owner, client, date range", "Slice the book", "E7-S3 / FR-DASH-07", "Not Started", 0, "", ""],
    ["S4", "S-7", "Ananya", "Sales", "Requirements", "Role-scoped default filter (salesOwner = me) + empty/error polish", "My book of business", "E3 / FR-REQ-09", "Not Started", 0, "", ""],
    ["S4", "T-M4", "Mohit", "TA", "Candidates", "Profile submitted / shortlist dates + interview round", "Interview tracking", "E4 / FR-CAN-04", "Not Started", 0, "", ""],
    ["S4", "T-M5", "Mohit", "TA", "Candidates", "Candidate detail page + pipeline age / candidate RAG", "Aging & risk", "E4 / FR-CAN-08", "Not Started", 0, "", ""],
    ["S4", "T-R2", "Rohit", "TA", "Candidates", "Select candidate (Selected = Yes) with business rules", "Selection gate", "E4-S4 / FR-CAN-05", "Not Started", 0, "", ""],
    ["S4", "A-6", "Harsh", "Admin", "Admin", "Audit log write on mutations + Admin audit viewer", "Who changed what", "E8-S1/S2 / FR-AUD", "Not Started", 0, "", ""],
    ["S4", "D-S3", "Saleena", "Dashboard", "Dashboard", "KPIs: pending handoff, pipeline, selected, offers, joined", "Funnel snapshot", "E7 / FR-DASH-02/03", "Not Started", 0, "", ""],
    ["S4", "D-A2", "Akash", "Dashboard", "Dashboard", "Debounced refetch + URL query sync (shareable filters)", "Shareable leadership view", "E7-S3 / FR-DASH-07", "Not Started", 0, "", ""],
    ["S4", "R-G3", "Gaurav", "Review", "Review", "Candidates + duplicates + select checklist vs FR-CAN", "TA module sign-off", "QA", "Not Started", 0, "", ""],
    ["S5", "T-M6", "Mohit", "TA", "Candidates", "Duplicate mobile/email warning banner (with confirm)", "No silent duplicates", "E4-S3 / FR-CAN-06/07", "Not Started", 0, "", ""],
    ["S5", "T-R3", "Rohit", "TA", "Offers", "Create offer from selected candidate (dates, CTC, DOJ, status)", "Offer initiated", "E5-S1 / FR-OFF-01", "Not Started", 0, "", ""],
    ["S5", "T-R4", "Rohit", "TA", "Offers", "Offer list + detail; status Initiated → Released", "Offer tracker", "E5-S2 / FR-OFF-02", "Not Started", 0, "", ""],
    ["S5", "H-S1", "Shalu", "HR", "Onboarding", "Onboarding list page", "Join pipeline visible", "E6 / FR-ONB", "Not Started", 0, "", ""],
    ["S5", "H-S2", "Shalu", "HR", "Onboarding", "Create onboarding from accepted offer + assign HR owner", "HR takes ownership", "E6-S1 / FR-ONB-01/02", "Not Started", 0, "", ""],
    ["S5", "A-H", "Harsh", "HR", "HR support", "Pair on onboarding RBAC / Joined position logic", "HR rules enforced server-side", "E6 / FR-ONB-06", "Not Started", 0, "", ""],
    ["S5", "D-S4", "Saleena", "Dashboard", "Dashboard", "Duplicate mobiles count card", "Data quality KPI", "E7 / FR-DASH-04", "Not Started", 0, "", ""],
    ["S5", "D-A3", "Akash", "Dashboard", "Dashboard", "Escalations / RED items list", "What needs attention today", "E7 / FR-DASH-09", "Not Started", 0, "", ""],
    ["S6", "T-R5", "Rohit", "TA", "Offers", "Offer TAT / RAG + block invalid/conflicting offers", "Offer SLA & rules", "E5-S3 / FR-OFF-04/06", "Not Started", 0, "", ""],
    ["S6", "T-R6", "Rohit", "TA", "Offers", "CTA: Accepted → Start onboarding (hand to HR)", "Clean TA → HR handoff", "E5 / FR-OFF-05", "Not Started", 0, "", ""],
    ["S6", "H-S3", "Shalu", "HR", "Onboarding", "Track docs pending, BGV status, joining formalities", "Pre-join checklist", "E6-S2 / FR-ONB-03", "Not Started", 0, "", ""],
    ["S6", "H-S4", "Shalu", "HR", "Onboarding", "Expected/actual DOJ + onboarding status flow", "Docs → BGV → formalities → Joined", "E6 / FR-ONB-04", "Not Started", 0, "", ""],
    ["S6", "A-7a", "Harsh", "Admin", "Admin", "CSV import draft + audit viewer polish", "Excel migration path (draft)", "E8 / FR-IMP", "Not Started", 0, "", ""],
    ["S6", "D-S5", "Saleena", "Dashboard", "Dashboard", "Candidate stage summary table", "Where candidates sit", "E7-S2 / FR-DASH-05", "Not Started", 0, "", ""],
    ["S6", "D-A4", "Akash", "Dashboard", "Dashboard", "Align API aggregations with Excel KPI definitions", "Parity with workbook", "E7 / FR-DASH", "Not Started", 0, "", ""],
    ["S6", "R-G4", "Gaurav", "Review", "Review", "Offer → onboarding → Joined E2E checklist", "End-to-end hire path", "QA", "Not Started", 0, "", ""],
    ["S7", "H-S5", "Shalu", "HR", "Onboarding", "Mark Joined → closed positions update on requirement", "Fill rate updates live", "E6-S3 / FR-ONB-04/06", "Not Started", 0, "", ""],
    ["S7", "H-S6", "Shalu", "HR", "Onboarding", "Onboarding TAT / RAG + cancel path if offer withdrawn", "Exception handling", "E6 / FR-ONB-05", "Not Started", 0, "", ""],
    ["S7", "A-7", "Harsh", "Admin", "Admin", "CSV import validate → commit (reqs/candidates)", "Excel migration path", "E8-S3 / FR-IMP-01/03", "Not Started", 0, "", ""],
    ["S7", "D-S6", "Saleena", "Dashboard", "Dashboard", "Requirement RAG summary + empty-state polish", "SLA health", "E7-S2 / FR-DASH-06", "Not Started", 0, "", ""],
    ["S7", "R-G5", "Gaurav", "Review", "Review", "Dashboard KPI vs Excel sample + regression pass", "Leadership confidence", "QA", "Not Started", 0, "", ""],
    ["S8", "ALL-1", "All owners", "All", "Hardening", "Bug bash on own module", "Quality pass", "M4", "Not Started", 0, "", ""],
    ["S8", "R-G5b", "Akash & Gaurav", "Review", "Review", "Full regression + final demo script", "Manager showcase ready", "QA", "Not Started", 0, "", ""],
    ["S8", "A-8", "Harsh", "Admin", "Admin", "Import final + audit polish", "Migration hardened", "E8", "Not Started", 0, "", ""],
    ["S8", "D-S7", "Saleena & Akash", "Dashboard", "Dashboard", "Empty/error polish on dashboard", "Leadership-ready UI", "E7", "Not Started", 0, "", ""],
    ["S8", "R-G6", "Gaurav", "Review", "Review", "Bug triage board (blocker / major / minor)", "Transparent quality status", "QA", "Not Started", 0, "", ""],
    ["Every", "R-A2", "Akash", "Review", "Review", "Smoke test critical path for merged PRs", "Nothing broken before demo", "QA", "Not Started", 0, "", ""],
]

SPRINT_BOARDS = [
    ["S1", "Foundations", "Login → create requirement", "Harsh: Login, seeds, masters | Ananya: Req list+create | Saleena: Dashboard shell | Review: role test users"],
    ["S2", "Sales + Admin usable", "Create user → client → req → assign TA", "Ananya: Detail/edit/TA assign | Harsh: Users/clients/JF | Mohit: Candidates list | Saleena: KPI layout"],
    ["S3", "TA pipeline starts", "One req with 2–3 candidates moving stages", "Mohit: Add candidate/stages | Rohit: Validation | Ananya: Status+RAG | Dashboard: KPIs+filters"],
    ["S4", "Selection quality", "Duplicate warning + select a candidate", "Mohit: Duplicates/detail | Rohit: Select | Harsh: Audit | Dashboard: funnel KPIs"],
    ["S5", "Offers", "Selected → offer released → accepted", "Rohit: Offer CRUD | Shalu: Onboarding create | Harsh: HR RBAC help | Dashboard: offers KPIs"],
    ["S6", "HR join path", "Accepted offer → Joined; positions closed", "Shalu: Docs/BGV/DOJ | Rohit: Offer RAG/CTA | Harsh: Import draft | Dashboard: stage/RAG"],
    ["S7", "Dashboard parity", "Filter dashboard; RED escalations; CSV import", "Saleena/Akash: Full KPIs | Shalu: Onboarding RAG | Harsh: CSV import | Review: KPI vs Excel"],
    ["S8", "Hardening & showcase", "E2E: Sales → TA → Offer → HR → Dashboard", "All: Bug bash | Review: Regression | Harsh: Import final | Dashboard: polish"],
]

HANDOFFS = [
    ["Admin → All", "Login works; masters API returns seeded values", "Harsh", "All module owners", "Open", ""],
    ["Sales → TA", "Requirement exists with REQ-#####, Active, TA assigned", "Ananya", "Mohit & Rohit", "Open", ""],
    ["TA → HR", "Candidate Selected; Offer Accepted", "Rohit", "Shalu", "Open", ""],
    ["HR → Dashboard", "Status Joined with actual DOJ; closed positions updated", "Shalu", "Saleena & Akash", "Open", ""],
    ["All → Review", "PR linked to Step ID + short test notes", "All owners", "Akash & Gaurav", "Open", ""],
]

CEREMONIES = [
    ["Planning", "Start of sprint", "Module owners pick Step IDs by dependency graph"],
    ["Daily", "Daily", "Blockers only"],
    ["Review", "End of sprint", "Akash leads demo script; owners demo"],
    ["Retro", "End of sprint", "Doc gaps + handoff friction"],
]

STATUSES = ["Not Started", "In Progress", "Blocked", "In Review", "Done"]
MODULES = ["Sales", "TA", "HR", "Admin", "Dashboard", "Review", "All"]
OWNERS = [
    "Ananya", "Mohit", "Rohit", "Shalu", "Harsh", "Saleena", "Akash", "Gaurav",
    "Saleena & Akash", "Akash & Gaurav", "Mohit & Rohit", "All owners",
]


def build():
    wb = Workbook()

    # ========== Dashboard ==========
    dash = wb.active
    dash.title = "Dashboard"
    dash.sheet_view.showGridLines = False
    set_widths(dash, {
        "A": 3, "B": 22, "C": 14, "D": 14, "E": 14, "F": 14, "G": 14, "H": 14, "I": 14, "J": 18, "K": 3, "L": 22, "M": 12, "N": 12,
    })

    dash.merge_cells("B2:J2")
    t = dash["B2"]
    t.value = "Service Staffing Tracker — Sprint Planning Tracker"
    t.font = Font(name="Calibri", bold=True, size=20, color=WHITE)
    t.fill = fill(NAVY)
    t.alignment = Alignment(horizontal="left", vertical="center")
    dash.row_dimensions[2].height = 36

    dash.merge_cells("B3:J3")
    s = dash["B3"]
    s.value = "MVP team tracker · Update Status & % Done on Sprint Backlog · Dashboard KPIs refresh automatically"
    s.font = font(size=10, color=TEAL)
    s.fill = fill(SOFT)
    s.alignment = Alignment(horizontal="left", vertical="center")
    dash.row_dimensions[3].height = 22

    # KPI cards
    kpi_labels = [
        ("B5", "C5", "Total Steps", '=COUNTA(\'Sprint Backlog\'!B7:B200)-COUNTBLANK(\'Sprint Backlog\'!B7:B200)'),
        ("D5", "E5", "Done", '=COUNTIF(\'Sprint Backlog\'!I7:I200,"Done")'),
        ("F5", "G5", "In Progress", '=COUNTIF(\'Sprint Backlog\'!I7:I200,"In Progress")'),
        ("H5", "I5", "Blocked", '=COUNTIF(\'Sprint Backlog\'!I7:I200,"Blocked")'),
        ("J5", "J5", "% Complete", '=IF(C6=0,0,E6/C6)'),
    ]
    # Simpler layout: label row 5, value row 6
    labels = ["Total Steps", "Done", "In Progress", "Blocked", "% Complete", "In Review"]
    formulas = [
        "=COUNTA('Sprint Backlog'!B7:B63)",
        "=COUNTIF('Sprint Backlog'!I7:I63,\"Done\")",
        "=COUNTIF('Sprint Backlog'!I7:I63,\"In Progress\")",
        "=COUNTIF('Sprint Backlog'!I7:I63,\"Blocked\")",
        "=IF(B6=0,0,C6/B6)",
        "=COUNTIF('Sprint Backlog'!I7:I63,\"In Review\")",
    ]
    cols = [2, 3, 4, 5, 6, 7]  # B..G
    for i, col in enumerate(cols):
        lc = dash.cell(row=5, column=col, value=labels[i])
        lc.fill = fill(TEAL)
        lc.font = font(bold=True, size=9, color=WHITE)
        lc.alignment = Alignment(horizontal="center", vertical="center")
        lc.border = thin
        vc = dash.cell(row=6, column=col)
        vc.value = formulas[i]
        vc.fill = fill(WHITE)
        vc.font = Font(name="Calibri", bold=True, size=18, color=NAVY)
        vc.alignment = Alignment(horizontal="center", vertical="center")
        vc.border = thick_outer
        dash.row_dimensions[6].height = 32
    dash["F6"].number_format = "0%"

    # Instructions panel
    dash.merge_cells("B8:G8")
    dash["B8"] = "How to use this workbook"
    dash["B8"].font = font(bold=True, size=12, color=WHITE)
    dash["B8"].fill = fill(NAVY)
    dash["B8"].alignment = Alignment(horizontal="left", vertical="center")

    instructions = [
        "1. Update Status and % Done on the Sprint Backlog sheet (dropdowns provided).",
        "2. Filter Backlog by Sprint, Owner, or Module using the table filters.",
        "3. Check By Person / By Sprint sheets for progress rollups (formula-driven).",
        "4. Mark Handoffs as Met when the contract criteria are satisfied.",
        "5. Use Sprint Boards for manager demo talking points each sprint review.",
        "6. Source of truth doc: docs/12-planning/TEAM_SPRINT_PLAN.md",
    ]
    for i, text in enumerate(instructions):
        r = 9 + i
        dash.merge_cells(f"B{r}:G{r}")
        cell = dash[f"B{r}"]
        cell.value = text
        cell.font = font(size=10)
        cell.fill = fill(LIGHT)
        cell.border = thin
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        dash.row_dimensions[r].height = 18

    # Milestone strip
    dash.merge_cells("B16:G16")
    dash["B16"] = "Milestones"
    dash["B16"].font = font(bold=True, size=12, color=WHITE)
    dash["B16"].fill = fill(TEAL)

    ms_headers = ["Milestone", "Sprints", "Theme", "Manager Demo"]
    for i, h in enumerate(ms_headers):
        c = dash.cell(row=17, column=2 + i, value=h)
        c.fill = fill(NAVY)
        c.font = font(bold=True, size=10, color=WHITE)
        c.border = thin
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    for i, row in enumerate(MILESTONES):
        for j, val in enumerate(row):
            c = dash.cell(row=18 + i, column=2 + j, value=val)
            c.border = thin
            c.font = font(size=10)
            c.alignment = Alignment(wrap_text=True, vertical="center")
            if i % 2 == 0:
                c.fill = fill(ROW_ALT)
        dash.row_dimensions[18 + i].height = 28

    # Chart helper (sprint progress)
    dash["L5"] = "Sprint"
    dash["M5"] = "Steps"
    dash["N5"] = "Done"
    for cell_ref in ("L5", "M5", "N5"):
        dash[cell_ref].fill = fill(NAVY)
        dash[cell_ref].font = font(bold=True, size=10, color=WHITE)
        dash[cell_ref].alignment = Alignment(horizontal="center")

    for i in range(1, 9):
        sprint = f"S{i}"
        r = 5 + i
        dash.cell(row=r, column=12, value=sprint)
        dash.cell(row=r, column=13).value = f"=COUNTIF('Sprint Backlog'!$A$7:$A$63,L{r})"
        dash.cell(row=r, column=14).value = f"=COUNTIFS('Sprint Backlog'!$A$7:$A$63,L{r},'Sprint Backlog'!$I$7:$I$63,\"Done\")"
        for c in range(12, 15):
            dash.cell(row=r, column=c).border = thin
            dash.cell(row=r, column=c).font = font(size=10)
            dash.cell(row=r, column=c).alignment = Alignment(horizontal="center")

    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.title = "Steps vs Done by Sprint"
    chart.y_axis.title = "Count"
    chart.x_axis.title = None
    chart.style = 10
    data = Reference(dash, min_col=13, min_row=5, max_col=14, max_row=13)
    cats = Reference(dash, min_col=12, min_row=6, max_row=13)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.shape = 4
    chart.width = 12
    chart.height = 8
    dash.add_chart(chart, "B23")

    dash.merge_cells("L15:N16")
    dash["L15"] = "Legend: update backlog Status to drive KPIs & chart"
    dash["L15"].font = font(size=9, color=SLATE)
    dash["L15"].alignment = Alignment(wrap_text=True, vertical="top")

    # ========== Ownership ==========
    own = wb.create_sheet("Ownership")
    own.sheet_view.showGridLines = False
    set_widths(own, {"A": 3, "B": 28, "C": 20, "D": 42, "E": 42, "F": 3})

    own.merge_cells("B2:E2")
    own["B2"] = "Module Ownership"
    own["B2"].font = Font(name="Calibri", bold=True, size=16, color=WHITE)
    own["B2"].fill = fill(NAVY)
    own["B2"].alignment = Alignment(horizontal="left", vertical="center")
    own.row_dimensions[2].height = 28

    headers = ["Module / Department", "Member(s)", "Primary Epics", "Demo Focus"]
    for i, h in enumerate(headers):
        c = own.cell(row=4, column=2 + i, value=h)
        c.fill = fill(TEAL)
        c.font = font(bold=True, size=11, color=WHITE)
        c.border = thin
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    own.row_dimensions[4].height = 22

    for i, row in enumerate(OWNERSHIP):
        for j, val in enumerate(row):
            c = own.cell(row=5 + i, column=2 + j, value=val)
            c.border = thin
            c.font = font(size=10)
            c.alignment = Alignment(wrap_text=True, vertical="center")
            if i % 2 == 0:
                c.fill = fill(ROW_ALT)
        own.row_dimensions[5 + i].height = 32

    own.merge_cells("B12:E12")
    own["B12"] = "Offer ownership split: Rohit (TA) create→release · Shalu (HR) accept→onboarding"
    own["B12"].font = font(size=10, color=ACCENT)
    own["B12"].fill = fill(LIGHT)
    own["B12"].alignment = Alignment(wrap_text=True, vertical="center")
    own.row_dimensions[12].height = 22

    own.merge_cells("B14:E14")
    own["B14"] = "Dependency: Admin unlocks all → Sales → TA → HR → Dashboard enrichment · Review every sprint"
    own["B14"].font = font(size=10, color=NAVY)
    own["B14"].fill = fill(SOFT)

    own.merge_cells("B16:E16")
    own["B16"] = "Ceremonies"
    own["B16"].font = font(bold=True, size=12, color=WHITE)
    own["B16"].fill = fill(NAVY)

    for i, h in enumerate(["Ceremony", "Cadence", "Owner / Focus"]):
        c = own.cell(row=17, column=2 + i, value=h)
        c.fill = fill(TEAL)
        c.font = font(bold=True, size=10, color=WHITE)
        c.border = thin
    for i, row in enumerate(CEREMONIES):
        for j, val in enumerate(row):
            c = own.cell(row=18 + i, column=2 + j, value=val)
            c.border = thin
            c.font = font(size=10)
            c.alignment = Alignment(wrap_text=True, vertical="center")

    # ========== Sprint Backlog ==========
    bl = wb.create_sheet("Sprint Backlog")
    bl.sheet_view.showGridLines = False
    set_widths(bl, {
        "A": 10, "B": 10, "C": 16, "D": 12, "E": 14, "F": 55, "G": 36, "H": 24,
        "I": 14, "J": 10, "K": 28, "L": 22,
    })

    bl.merge_cells("A1:L1")
    bl["A1"] = "Sprint Backlog — Track every step (edit Status / % Done / Notes / PR Link)"
    bl["A1"].font = Font(name="Calibri", bold=True, size=14, color=WHITE)
    bl["A1"].fill = fill(NAVY)
    bl["A1"].alignment = Alignment(horizontal="left", vertical="center")
    bl.row_dimensions[1].height = 26

    bl.merge_cells("A2:L2")
    bl["A2"] = "Done when: UI + API (where applicable) + short test note · Link PRs to Step ID"
    bl["A2"].font = font(size=9, color=TEAL)
    bl["A2"].fill = fill(SOFT)

    bl_headers = [
        "Sprint", "Step ID", "Owner", "Module", "Track", "Deliverable",
        "Demo Line", "Epic / FR", "Status", "% Done", "Notes", "PR Link",
    ]
    for i, h in enumerate(bl_headers):
        c = bl.cell(row=6, column=1 + i, value=h)
        c.fill = fill(NAVY)
        c.font = font(bold=True, size=10, color=WHITE)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = thin
    bl.row_dimensions[6].height = 28

    for i, row in enumerate(STEPS):
        r = 7 + i
        for j, val in enumerate(row):
            c = bl.cell(row=r, column=1 + j, value=val)
            c.border = thin
            c.font = font(size=9)
            c.alignment = Alignment(vertical="center", wrap_text=True, horizontal="left")
            if j in (0, 1, 3, 8, 9):
                c.alignment = Alignment(vertical="center", wrap_text=True, horizontal="center")
            if j == 9:
                c.number_format = "0%"
                # store as fraction
                c.value = float(val) if isinstance(val, (int, float)) else 0
            if i % 2 == 1:
                c.fill = fill(ROW_ALT)
        bl.row_dimensions[r].height = 36

    last_row = 6 + len(STEPS)
    table = Table(displayName="SprintBacklog", ref=f"A6:L{last_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=False,
    )
    bl.add_table(table)

    # Data validations
    dv_status = DataValidation(type="list", formula1='"' + ",".join(STATUSES) + '"', allow_blank=True)
    dv_status.error = "Pick a valid status"
    dv_status.errorTitle = "Invalid Status"
    bl.add_data_validation(dv_status)
    dv_status.add(f"I7:I{last_row}")

    dv_owner = DataValidation(type="list", formula1='"' + ",".join(OWNERS) + '"', allow_blank=True)
    bl.add_data_validation(dv_owner)
    dv_owner.add(f"C7:C{last_row}")

    dv_module = DataValidation(type="list", formula1='"' + ",".join(MODULES) + '"', allow_blank=True)
    bl.add_data_validation(dv_module)
    dv_module.add(f"D7:D{last_row}")

    # Conditional formatting for Status
    status_colors = {
        "Done": GREEN,
        "In Progress": TEAL,
        "Blocked": RED,
        "In Review": AMBER,
        "Not Started": SLATE,
    }
    for status, color in status_colors.items():
        bl.conditional_formatting.add(
            f"I7:I{last_row}",
            FormulaRule(
                formula=[f'$I7="{status}"'],
                fill=fill(color),
                font=Font(name="Calibri", bold=True, size=9, color=WHITE),
            ),
        )

    bl.freeze_panes = "A7"

    # ========== By Sprint ==========
    bs = wb.create_sheet("By Sprint")
    bs.sheet_view.showGridLines = False
    set_widths(bs, {"A": 3, "B": 10, "C": 22, "D": 48, "E": 70, "F": 12, "G": 12, "H": 12, "I": 12})

    bs.merge_cells("B2:I2")
    bs["B2"] = "Sprint Boards — Focus & Manager Demo"
    bs["B2"].font = Font(name="Calibri", bold=True, size=16, color=WHITE)
    bs["B2"].fill = fill(NAVY)
    bs["B2"].alignment = Alignment(vertical="center")
    bs.row_dimensions[2].height = 28

    sb_headers = ["Sprint", "Theme", "Manager Demo", "Who Does What", "Total", "Done", "In Progress", "% Complete"]
    for i, h in enumerate(sb_headers):
        c = bs.cell(row=4, column=2 + i, value=h)
        c.fill = fill(TEAL)
        c.font = font(bold=True, size=10, color=WHITE)
        c.border = thin
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    bs.row_dimensions[4].height = 30

    for i, row in enumerate(SPRINT_BOARDS):
        r = 5 + i
        sprint = row[0]
        for j, val in enumerate(row):
            c = bs.cell(row=r, column=2 + j, value=val)
            c.border = thin
            c.font = font(size=9)
            c.alignment = Alignment(wrap_text=True, vertical="center")
        # formula columns F G H I = cols 6,7,8,9
        bs.cell(row=r, column=6).value = f"=COUNTIF('Sprint Backlog'!$A$7:$A$63,B{r})"
        bs.cell(row=r, column=7).value = f"=COUNTIFS('Sprint Backlog'!$A$7:$A$63,B{r},'Sprint Backlog'!$I$7:$I$63,\"Done\")"
        bs.cell(row=r, column=8).value = f"=COUNTIFS('Sprint Backlog'!$A$7:$A$63,B{r},'Sprint Backlog'!$I$7:$I$63,\"In Progress\")"
        bs.cell(row=r, column=9).value = f"=IF(F{r}=0,0,G{r}/F{r})"
        bs.cell(row=r, column=9).number_format = "0%"
        for col in range(6, 10):
            c = bs.cell(row=r, column=col)
            c.border = thin
            c.font = Font(name="Calibri", bold=True, size=11, color=NAVY)
            c.alignment = Alignment(horizontal="center", vertical="center")
        bs.row_dimensions[r].height = 48
        if i % 2 == 0:
            for col in range(2, 6):
                bs.cell(row=r, column=col).fill = fill(ROW_ALT)

    # ========== By Person ==========
    bp = wb.create_sheet("By Person")
    bp.sheet_view.showGridLines = False
    set_widths(bp, {"A": 3, "B": 18, "C": 16, "D": 12, "E": 12, "F": 12, "G": 12, "H": 14})

    bp.merge_cells("B2:H2")
    bp["B2"] = "Progress by Person"
    bp["B2"].font = Font(name="Calibri", bold=True, size=16, color=WHITE)
    bp["B2"].fill = fill(NAVY)
    bp["B2"].alignment = Alignment(vertical="center")
    bp.row_dimensions[2].height = 28

    people = [
        ("Ananya", "Sales"),
        ("Mohit", "TA"),
        ("Rohit", "TA"),
        ("Shalu", "HR"),
        ("Harsh", "Admin / HR"),
        ("Saleena", "Dashboard"),
        ("Akash", "Dashboard / Review"),
        ("Gaurav", "Review"),
    ]
    for i, h in enumerate(["Person", "Primary Module", "Total Steps", "Done", "In Progress", "Blocked", "% Complete"]):
        c = bp.cell(row=4, column=2 + i, value=h)
        c.fill = fill(TEAL)
        c.font = font(bold=True, size=10, color=WHITE)
        c.border = thin
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    for i, (person, module) in enumerate(people):
        r = 5 + i
        bp.cell(row=r, column=2, value=person)
        bp.cell(row=r, column=3, value=module)
        # COUNTIF on owner column — also catch shared names partially via COUNTIF exact
        bp.cell(row=r, column=4).value = f"=COUNTIF('Sprint Backlog'!$C$7:$C$63,B{r})"
        bp.cell(row=r, column=5).value = f"=COUNTIFS('Sprint Backlog'!$C$7:$C$63,B{r},'Sprint Backlog'!$I$7:$I$63,\"Done\")"
        bp.cell(row=r, column=6).value = f"=COUNTIFS('Sprint Backlog'!$C$7:$C$63,B{r},'Sprint Backlog'!$I$7:$I$63,\"In Progress\")"
        bp.cell(row=r, column=7).value = f"=COUNTIFS('Sprint Backlog'!$C$7:$C$63,B{r},'Sprint Backlog'!$I$7:$I$63,\"Blocked\")"
        bp.cell(row=r, column=8).value = f"=IF(D{r}=0,0,E{r}/D{r})"
        bp.cell(row=r, column=8).number_format = "0%"
        for col in range(2, 9):
            c = bp.cell(row=r, column=col)
            c.border = thin
            c.font = font(size=10)
            c.alignment = Alignment(horizontal="center" if col > 3 else "left", vertical="center")
            if i % 2 == 0:
                c.fill = fill(ROW_ALT)
        bp.row_dimensions[r].height = 22

    bp.merge_cells("B14:H14")
    bp["B14"] = "Note: Shared owners (e.g. Saleena & Akash) are counted on those exact Owner cell values in Sprint Backlog."
    bp["B14"].font = font(size=9, color=SLATE)

    # ========== Handoffs ==========
    ho = wb.create_sheet("Handoffs")
    ho.sheet_view.showGridLines = False
    set_widths(ho, {"A": 3, "B": 18, "C": 55, "D": 16, "E": 18, "F": 12, "G": 28})

    ho.merge_cells("B2:G2")
    ho["B2"] = "Handoff Contracts"
    ho["B2"].font = Font(name="Calibri", bold=True, size=16, color=WHITE)
    ho["B2"].fill = fill(NAVY)
    ho["B2"].alignment = Alignment(vertical="center")
    ho.row_dimensions[2].height = 28

    for i, h in enumerate(["From → To", "Done When (Contract)", "From Owner", "To Owner", "Status", "Notes"]):
        c = ho.cell(row=4, column=2 + i, value=h)
        c.fill = fill(TEAL)
        c.font = font(bold=True, size=10, color=WHITE)
        c.border = thin
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    for i, row in enumerate(HANDOFFS):
        r = 5 + i
        for j, val in enumerate(row):
            c = ho.cell(row=r, column=2 + j, value=val)
            c.border = thin
            c.font = font(size=10)
            c.alignment = Alignment(wrap_text=True, vertical="center")
            if i % 2 == 0:
                c.fill = fill(ROW_ALT)
        ho.row_dimensions[r].height = 36

    dv_ho = DataValidation(type="list", formula1='"Open,In Progress,Met,Blocked"', allow_blank=True)
    ho.add_data_validation(dv_ho)
    dv_ho.add("F5:F9")

    for status, color in [("Met", GREEN), ("Blocked", RED), ("In Progress", TEAL), ("Open", SLATE)]:
        ho.conditional_formatting.add(
            "F5:F9",
            FormulaRule(
                formula=[f'$F5="{status}"'],
                fill=fill(color),
                font=Font(name="Calibri", bold=True, size=9, color=WHITE),
            ),
        )

    # ========== Manager Demo ==========
    md = wb.create_sheet("Manager Demo")
    md.sheet_view.showGridLines = False
    set_widths(md, {"A": 3, "B": 10, "C": 55, "D": 14, "E": 40})

    md.merge_cells("B2:E2")
    md["B2"] = "Manager Showcase Checklist (use every sprint review ~10 min)"
    md["B2"].font = Font(name="Calibri", bold=True, size=14, color=WHITE)
    md["B2"].fill = fill(NAVY)
    md["B2"].alignment = Alignment(vertical="center")
    md.row_dimensions[2].height = 28

    agenda = [
        ["1", "Who owns what — ownership table", "2 min", "Show Ownership sheet"],
        ["2", "Steps completed — checkboxes by Step ID", "2 min", "Filter Sprint Backlog by current sprint + Done"],
        ["3", "Live demo — one vertical slice only", "4 min", "See Sprint Boards → Manager Demo column"],
        ["4", "What's next / dependencies", "1 min", "Handoffs sheet + next sprint focus"],
        ["5", "Risks — blockers Review found", "1 min", "Filter Status = Blocked on Sprint Backlog"],
    ]
    for i, h in enumerate(["#", "Agenda Item", "Time", "Where in this workbook"]):
        c = md.cell(row=4, column=2 + i, value=h)
        c.fill = fill(TEAL)
        c.font = font(bold=True, size=10, color=WHITE)
        c.border = thin

    for i, row in enumerate(agenda):
        r = 5 + i
        for j, val in enumerate(row):
            c = md.cell(row=r, column=2 + j, value=val)
            c.border = thin
            c.font = font(size=10)
            c.alignment = Alignment(wrap_text=True, vertical="center")
            if i % 2 == 0:
                c.fill = fill(ROW_ALT)
        md.row_dimensions[r].height = 24

    md.merge_cells("B11:E11")
    md["B11"] = "Per-sprint demo lines (copy into slides)"
    md["B11"].font = font(bold=True, size=12, color=WHITE)
    md["B11"].fill = fill(NAVY)

    for i, h in enumerate(["Sprint", "Theme", "Demo Line"]):
        c = md.cell(row=12, column=2 + i, value=h)
        c.fill = fill(TEAL)
        c.font = font(bold=True, size=10, color=WHITE)
        c.border = thin

    for i, row in enumerate(SPRINT_BOARDS):
        r = 13 + i
        vals = [row[0], row[1], row[2]]
        for j, val in enumerate(vals):
            c = md.cell(row=r, column=2 + j, value=val)
            c.border = thin
            c.font = font(size=10)
            c.alignment = Alignment(wrap_text=True, vertical="center")
            if i % 2 == 0:
                c.fill = fill(ROW_ALT)
        md.row_dimensions[r].height = 28

    md.merge_cells("B22:E22")
    md["B22"] = "References: docs/12-planning/TEAM_SPRINT_PLAN.md · EPICS_AND_STORIES.md · SPRINT_AND_MILESTONES.md"
    md["B22"].font = font(size=9, color=SLATE)

    # ========== Lookups (hidden helper) ==========
    lu = wb.create_sheet("Lookups")
    for i, s in enumerate(STATUSES):
        lu.cell(row=1 + i, column=1, value=s)
    for i, o in enumerate(OWNERS):
        lu.cell(row=1 + i, column=2, value=o)
    for i, m in enumerate(MODULES):
        lu.cell(row=1 + i, column=3, value=m)
    lu.sheet_state = "hidden"

    # Print / freeze niceties
    dash.freeze_panes = "B5"
    own.freeze_panes = "B5"

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT}")
    print(f"Steps: {len(STEPS)}")


if __name__ == "__main__":
    build()
